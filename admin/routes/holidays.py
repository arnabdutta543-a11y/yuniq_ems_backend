import csv
import io
import datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, status
from sqlalchemy.orm import Session
from typing import List, Optional
from openpyxl import load_workbook
from admin.database import get_db
from admin import models, schemas

router = APIRouter(prefix="/holidays", tags=["Admin Holidays"])

@router.get("/", response_model=List[schemas.HolidayOut])
def get_holidays(office: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(models.Holiday)
    if office:
        query = query.filter(models.Holiday.office.in_([office, "All"]))
    return query.order_by(models.Holiday.date.asc()).all()

@router.post("/", response_model=schemas.HolidayOut)
def create_holiday(holiday_in: schemas.HolidayCreate, db: Session = Depends(get_db)):
    day_name = holiday_in.date.strftime("%a")
    db_holiday = models.Holiday(
        name=holiday_in.name, date=holiday_in.date, day=day_name, office=holiday_in.office
    )
    db.add(db_holiday)
    db.commit()
    db.refresh(db_holiday)
    return db_holiday

@router.delete("/{id}")
def delete_holiday(id: int, db: Session = Depends(get_db)):
    holiday = db.query(models.Holiday).filter(models.Holiday.id == id).first()
    if not holiday:
        raise HTTPException(status_code=404, detail="Holiday not found.")
    db.delete(holiday)
    db.commit()
    return {"message": "Holiday successfully deleted."}

@router.post("/upload")
def upload_holidays(
    file: UploadFile = File(...),
    office: str = Form("All"),
    db: Session = Depends(get_db)
):
    """
    Supports Excel (.xlsx) and CSV spreadsheet imports of holidays.
    Expected Columns: Holiday Name, Date (YYYY-MM-DD), Office (optional)
    """
    filename = file.filename.lower()
    holidays_to_insert = []

    try:
        if filename.endswith(".xlsx"):
            contents = file.file.read()
            wb = load_workbook(io.BytesIO(contents))
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                h_name = str(row[0]).strip()
                h_date = row[1]
                h_office = str(row[2]).strip() if len(row) > 2 and row[2] else office

                if isinstance(h_date, datetime.datetime):
                    parsed_date = h_date.date()
                elif isinstance(h_date, datetime.date):
                    parsed_date = h_date
                else:
                    try:
                        parsed_date = datetime.datetime.strptime(str(h_date).strip()[:10], "%Y-%m-%d").date()
                    except ValueError:
                        try:
                            parsed_date = datetime.datetime.strptime(str(h_date).strip()[:10], "%d-%m-%Y").date()
                        except ValueError:
                            continue

                day_name = parsed_date.strftime("%a")
                holidays_to_insert.append(models.Holiday(
                    name=h_name, date=parsed_date, day=day_name, office=h_office
                ))

        elif filename.endswith(".csv"):
            contents = file.file.read().decode("utf-8")
            csv_reader = csv.reader(io.StringIO(contents))
            next(csv_reader, None)  # skip header

            for row in csv_reader:
                if not row or not row[0]:
                    continue
                h_name = row[0].strip()
                h_date_str = row[1].strip()
                h_office = row[2].strip() if len(row) > 2 and row[2] else office

                try:
                    parsed_date = datetime.datetime.strptime(h_date_str[:10], "%Y-%m-%d").date()
                except ValueError:
                    try:
                        parsed_date = datetime.datetime.strptime(h_date_str[:10], "%d-%m-%Y").date()
                    except ValueError:
                        continue

                day_name = parsed_date.strftime("%a")
                holidays_to_insert.append(models.Holiday(
                    name=h_name, date=parsed_date, day=day_name, office=h_office
                ))
        else:
            raise HTTPException(status_code=400, detail="Invalid file format. Please upload an Excel (.xlsx) or CSV file.")

        if not holidays_to_insert:
            raise HTTPException(status_code=400, detail="No valid holiday records found in the uploaded file.")

        db.add_all(holidays_to_insert)
        db.commit()

        return {
            "message": f"Successfully imported {len(holidays_to_insert)} holidays!",
            "count": len(holidays_to_insert)
        }

    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error parsing uploaded file: {str(e)}"
        )
